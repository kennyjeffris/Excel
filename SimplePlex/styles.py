from openpyxl.styles import Alignment, PatternFill, colors
from openpyxl.styles.borders import Border, Side


def get():
    useful_styles = {}
    useful_styles['medium'] = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

    useful_styles['thin'] = Border(left=Side(style='thin'),
              right=Side(style='thin'),
              top=Side(style='thin'),
              bottom=Side(style='thin'))

    useful_styles['medium_thin'] = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    useful_styles['center_center'] = Alignment(horizontal='center', vertical='center')

    useful_styles['right_center'] = Alignment(horizontal='right', vertical='center')
    # Colors
    useful_styles['light_blue'] = '95B3D7'
    useful_styles['red'] = 'FF5050'
    # Highlight colors
    useful_styles['yellow_fill'] = PatternFill('solid', fgColor=colors.YELLOW)
    useful_styles['red_fill'] = PatternFill('solid', fgColor=red)
    useful_styles['light_blue_fill'] = PatternFill('solid', fgColor=light_blue)

    return useful_styles