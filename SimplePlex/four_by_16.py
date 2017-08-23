import styles
from tkinter import messagebox
from openpyxl.utils import get_column_letter
from sys import exit
from func import poly_fit, as_text, prep_lists
from openpyxl.chart import ScatterChart, Reference, Series, marker
from openpyxl.drawing.fill import ColorChoice
###########################################


def format(wb, analytes, max_row, max_col):
    global style
    style = styles.get()
    headerlist, searchlist = prep_lists(analytes)
    wb = summaries_1_2(wb, analytes, headerlist, searchlist, max_row, max_col)
    wb = summary_3(wb, analytes, headerlist, searchlist, max_row, max_col)
    return wb


def get_items(sheet, analyte, data, max_row, max_col):
    """Get the items in a column of a request sheet."""
    try:
        done = False
        count = 1
        while not done:
            letter = get_column_letter(count)
            ind = letter + '1'
            # Check for instances where variable names change and a list of options are available.
            if isinstance(data, (list, tuple)):
                iter_length = len(data)
            else:
                iter_length = 1
            for x in range(0, iter_length):
                if iter_length == 1:
                    obj = data
                else:
                    obj = data[x]

                if obj == sheet[ind].value:
                    item = []
                    for l in range(analyte, max_row, 4):
                        ind = letter + str(l + 1)
                        cell = sheet[ind]
                        try:
                            if cell.value.isspace() or cell.value == '' or cell.value is None or cell.value == "NaN":
                                item.append("ND")
                            else:
                                item.append(float(cell.value))
                        except Exception:
                            item.append(cell.value)
                    return item
            count += 1
            if count > max_col:
                raise ValueError('Item not found')
    except ValueError as error:
        messagebox.showerror(message="Missing item {}.  Please export your data with "
                             "this item included".format(data), title="Failure")
        exit()


def summaries_1_2(wb, analytes, headerList, searchList, max_row, max_col):
    ws1 = wb.worksheets[0]
    ws2 = wb.create_sheet(title='Summary 1')
    ws3 = wb.create_sheet(title='Summary 2')
    for page in range(1, 3):
        if page == 1:
            workingSheet = ws2
            working_headerList = headerList[0]
            working_searchList = searchList[0]
        else:
            workingSheet = ws3
            working_headerList = headerList[1]
            working_searchList = searchList[1]
        for i in range(0, 4):
            startRow = i * 20 + 1
            analyteString = 'Analyte {} ({})'.format(i + 1, analytes[i])
            workingSheet['A{}'.format(startRow)] = analyteString
            if page == 1:
                workingSheet.merge_cells(start_row=startRow + 1, start_column=3, end_row=startRow + 1, end_column=8)
                cell = workingSheet['C{}'.format(startRow + 1)]
                cell.value = 'RFU'
                cell.alignment = style['center_center']
                cell.fill = style['yellow_fill']

                workingSheet.merge_cells(start_row=startRow + 1, start_column=9, end_row=startRow + 1, end_column=12)
                cell = workingSheet['I{}'.format(startRow + 1)]
                cell.value = 'RFU-Bkgd'
                cell.alignment = style['center_center']
                cell.fill = style['yellow_fill']

                workingSheet.merge_cells(start_row=startRow + 1, start_column=13, end_row=startRow + 1, end_column=17)
                cell = workingSheet['M{}'.format(startRow + 1)]
                cell.value = 'Calculated Concentration'
                cell.alignment = style['center_center']
                cell.fill = style['yellow_fill']
                for index, col in enumerate(iterable=workingSheet.iter_cols(min_row=startRow + 1, max_row=startRow + 1,
                                                                            min_col=3, max_col=17)):
                    for cell in col:
                        cell.border = style['medium_thin']
            else:
                workingSheet.merge_cells(start_row=startRow + 1, start_column=3, end_row=startRow + 1, end_column=7)
                cell = workingSheet['C{}'.format(startRow + 1)]
                cell.value = 'Calculated Concentration'
                cell.alignment = style['center_center']
                cell.fill = style['yellow_fill']
                for index, col in enumerate(iterable=workingSheet.iter_cols(min_row=startRow + 1, max_row=startRow + 1,
                                                                            min_col=3, max_col=7)):
                    for cell in col:
                        cell.border = style['medium_thin']
            for index, col in enumerate(iterable=workingSheet.iter_cols(
                    min_row=startRow + 2,
                    min_col=1, max_row=startRow + 2,
                    max_col=len(working_headerList))):
                for cell in col:
                    cell.value = working_headerList[index]
                    cell.border = style['medium']
                    if i == 0:
                        length = len(as_text(cell.value)) + 2
                        workingSheet.column_dimensions[cell.column].width = length

            for index, row in enumerate(iterable=workingSheet.iter_rows(
                    min_row=startRow + 3,
                    max_col=1, max_row=startRow + 18)):
                for cell in row:
                    cell.value = index + 1
                    cell.border = style['medium_thin']

            for index, col in enumerate(iterable=workingSheet.iter_cols(
                    min_col=2, max_col=len(working_headerList))):
                feature = working_searchList[index]
                values = get_items(ws1, i + 1, feature, max_row, max_col)
                for index2, row in enumerate(iterable=workingSheet.iter_rows(
                        min_col=index + 2,
                        max_col=index + 2,
                        min_row=startRow + 3,
                        max_row=startRow + 18)):
                    for cell in row:
                        cell.value = values[index2]
                        if cell.value == 'ND':
                            cell.fill = style['red_fill']
                        cell.alignment = style['right_center']
                        cell.border = style['thin']
    return wb


def summary_3(wb, analytes, headerLists, searchLists, max_row, max_col):
    ws4 = wb.create_sheet(title='Summary 3')
    ws1 = wb.worksheets[0]
    headerList3 = headerLists[2]
    headerList4 = headerLists[3]
    headerList5 = headerLists[4]
    ws4['A1'].value = headerList4[0]
    for index, col in enumerate(iterable=ws4.iter_cols(
            min_row=2,
            min_col=1, max_row=2,
            max_col=len(headerList3))):
        for cell in col:
            cell.value = headerList3[index]
            cell.border = style['medium']

    for index, row in enumerate(iterable=ws4.iter_rows(
            min_row=3,
            max_col=1, max_row=18)):
        for cell in row:
            cell.value = index + 1
            cell.border = style['medium_thin']

    for index, col in enumerate(iterable=ws4.iter_cols(
            min_col=2, max_col=len(headerList3),
            min_row=3, max_row=18)):

        values = get_items(ws1, index + 1, headerList4[0], max_row, max_col)
        for index2, row in enumerate(iterable=ws4.iter_rows(
                min_col=index + 2,
                max_col=index + 2,
                min_row=3,
                max_row=18)):
            for cell in row:
                cell.value = values[index2]
                if cell.value == 'ND':
                    cell.fill = style['red_fill']
                cell.alignment = style['right_center']
                cell.border = style['thin']

    ws4['A20'].value = headerList5[0]
    ws4['A21'].border = style['medium']
    for index, col in enumerate(iterable=ws4.iter_cols(min_row=21,
                                                       min_col=2, max_row=21, max_col=5)):
        for cell in col:
            cell.value = headerList3[index + 1]
            cell.border = style['medium']

    for index, row in enumerate(iterable=ws4.iter_rows(min_row=22, max_row=26,
                                                       min_col=1, max_col=1)):
        for cell in row:
            cell.value = headerList5[index + 1]
            cell.border = style['medium_thin']

    for index, col in enumerate(iterable=ws4.iter_cols(
            min_col=2, max_col=len(headerList3),
            min_row=22, max_row=26)):

        for index2, row in enumerate(iterable=ws4.iter_rows(
                min_col=index + 2,
                min_row=22,
                max_row=26)):
            values = get_items(ws1, index + 1, headerList4[index2 + 1], max_row, max_col)
            for cell in row:
                cell.value = values[0]
                cell.border = style['thin']

    ##############################################
    # Add plots to Summary 3

    for index, col in enumerate(iterable=ws4.iter_cols(
            min_col=2,
            max_col=5)):
        coeffs = []
        let = get_column_letter(2 + index)
        for index2, row in enumerate(iterable=ws4.iter_rows(
                range_string='{}22:{}26'.format(let,
                                                let))):
            for cell in row:
                coeffs.append(cell.value)
        xvalues = []
        values = []
        xvalues.extend((.1, 1, 10, 100, 1000, 10000))
        for num in range(0, len(xvalues)):
            if isinstance(xvalues[num], (int, float, complex)):
                values.append(poly_fit(xvalues[num], coeffs))
        for index3, row in enumerate(iterable=ws4.iter_rows(
                min_row=3,
                max_row=18,
                min_col=2 + index,
                max_col=2 + index)):
            for cell in row:
                xval = cell.value
                if isinstance(xval, (int, float, complex)):
                    xvalues.append(xval)
                    values.append(poly_fit(xval, coeffs))

        xvalues, values = zip(*sorted(zip(xvalues, values)))
        xvalues, values = (list(t) for t in zip(*sorted(zip(xvalues, values))))
        count1 = 0
        count2 = 0
        for index4, row in enumerate(iterable=ws4.iter_rows(
                range_string='{}28:{}72'.format(let, let))):
            for cell in row:
                if count1 < len(xvalues):
                    cell.value = xvalues[count1]
                    count1 += 1
                elif count1 >= len(values) and count2 < len(values):
                    cell.value = values[count2]
                    count2 += 1
                else:
                    break
        xref = Reference(ws4, min_col=2 + index, max_col=2 + index,
                         min_row=28, max_row=27 + len(xvalues))
        yref = Reference(ws4, min_col=2 + index, max_col=2 + index,
                         min_row=28 + len(xvalues),
                         max_row=27 + len(xvalues) + len(values))

        # Chart formatting
        chart = ScatterChart()
        # Title
        chart.title = '{}'.format(headerList3[index + 1])
        chart.style = 13
        chart.legend = None
        chart.x_axis.scaling.logBase = 10
        chart.y_axis.scaling.logBase = 10
        chart.x_axis.scaling.min = 0.1
        chart.y_axis.scaling.min = 0.1
        chart.x_axis.scaling.max = 10000
        chart.y_axis.scaling.max = 10000
        chart.y_axis.crossesAt = 0.1
        chart.y_axis.title = 'RFU'
        chart.x_axis.title = 'Concentration (pg/ml)'
        chart.x_axis.tickLblPos = "low"
        series = Series(yref, xref, title_from_data=False)
        series.marker = marker.Marker('circle')
        series.smooth = True
        series.graphicalProperties.line.width = 20000
        series.graphicalProperties.line.solidFill = ColorChoice(prstClr="orange")
        series.marker.graphicalProperties.solidFill = ColorChoice(prstClr="orange")  # Marker filling
        series.marker.graphicalProperties.line.solidFill = ColorChoice(prstClr="orange")  # Marker outline
        chart.series.append(series)
        ws4.add_chart(chart, 'H{}'.format((index * 15) + 1))
    return wb