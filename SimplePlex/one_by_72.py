import styles
from tkinter import messagebox
from openpyxl.utils import get_column_letter
from sys import exit
from func import poly_fit, as_text, prep_lists
from openpyxl.chart import ScatterChart, Reference, Series, marker
from openpyxl.drawing.fill import ColorChoice

def format(wb, analytes, max_row, max_col):
    global style
    style = styles.get()
    headerlist, searchlist = prep_lists(analytes)
    wb = summaries_1_2(wb, analytes, headerlist, searchlist, max_row, max_col)
    #wb = summary_3(wb, analytes, headerlist, searchlist, max_row, max_col)
    return wb


def get_items(sheet, data, max_row, max_col):
    """Get the items in a column of a request sheet."""
    try:
        done = False
        count = 1
        while not done:
            letter = get_column_letter(count)
            ind = letter + '1'
            # Check for instances where variable names change and a list of options are available.
            if isinstance(data, (list, tuple)):
                iter_length = len(data)
            else:
                iter_length = 1
            for x in range(0, iter_length):
                if iter_length == 1:
                    obj = data
                else:
                    obj = data[x]
                if obj == sheet[ind].value:
                    item = []
                    for l in range(1, max_row):
                        ind = letter + str(l + 1)
                        cell = sheet[ind]
                        try:
                            if cell.value.isspace() or cell.value == '' or cell.value is None or cell.value == "NaN":
                                item.append("ND")
                            else:
                                item.append(float(cell.value))
                        except Exception:
                            item.append(cell.value)
                    return item
            count += 1
            if count > max_col:
                raise ValueError('Item not found')
    except ValueError as error:
        messagebox.showerror(message="Missing item {}.  Please export your data with "
                             "this item included".format(data), title="Failure")
        exit()

def summaries_1_2(wb, analytes, headerList, searchList, max_row, max_col):
    ws1 = wb.worksheets[0]
    ws2 = wb.create_sheet(title='Summary 1')
    ws3 = wb.create_sheet(title='Summary 2')
    for page in range(1, 3):
        if page == 1:
            workingSheet = ws2
            working_headerList = headerList[0]
            working_searchList = searchList[0]
        else:
            workingSheet = ws3
            working_headerList = headerList[1]
            working_searchList = searchList[1]
        startRow = 1
        analyteString = analytes[0]
        workingSheet['A{}'.format(startRow)] = analyteString
        if page == 1:
            workingSheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=8)
            cell = workingSheet['C{}'.format(2)]
            cell.value = 'RFU'
            cell.alignment = style['center_center']
            cell.fill = style['yellow_fill']

            workingSheet.merge_cells(start_row=2, start_column=9, end_row=2, end_column=12)
            cell = workingSheet['I{}'.format(2)]
            cell.value = 'RFU-Bkgd'
            cell.alignment = style['center_center']
            cell.fill = style['yellow_fill']

            workingSheet.merge_cells(start_row=2, start_column=13, end_row=2, end_column=17)
            cell = workingSheet['M{}'.format(2)]
            cell.value = 'Calculated Concentration'
            cell.alignment = style['center_center']
            cell.fill = style['yellow_fill']
        for index, col in enumerate(iterable=workingSheet.iter_cols(min_row=2, max_row=2,
                                                                    min_col=3, max_col=17)):
            for cell in col:
                cell.border = style['medium_thin']
        else:
            for index, col in enumerate(iterable=workingSheet.iter_cols(min_row=2, max_row=2,
                                                                        min_col=3, max_col=7)):
                for cell in col:
                    cell.border = style['medium_thin']
            workingSheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=7)
            cell = workingSheet['C{}'.format(2)]
            cell.value = 'Calculated Concentration'
            cell.alignment = style['center_center']
            cell.fill = style['yellow_fill']
        for index, col in enumerate(iterable=workingSheet.iter_cols(
                min_row=3,
                min_col=1, max_row=3,
                max_col=len(working_headerList))):
            for cell in col:
                cell.value = working_headerList[index]
                cell.border = style['medium']
                length = len(as_text(cell.value)) + 9
                workingSheet.column_dimensions[cell.column].width = length

        for index, row in enumerate(iterable=workingSheet.iter_rows(
                min_row=4,
                max_col=1, max_row=75)):
            for cell in row:
                cell.value = index + 1
                cell.border = style['medium_thin']

        for index, col in enumerate(iterable=workingSheet.iter_cols(
                min_col=2, max_col=len(working_headerList))):
            feature = working_searchList[index]
            values = get_items(ws1, feature, max_row, max_col)
            for index2, row in enumerate(iterable=workingSheet.iter_rows(
                    min_col=index + 2,
                    max_col=index + 2,
                    min_row=startRow + 3,
                    max_row=75)):
                for cell in row:
                    cell.value = values[index2]
                    if cell.value == 'ND':
                        cell.fill = style['red_fill']
                    cell.alignment = style['right_center']
                    cell.border = style['thin']
    return wb

def summary_3(wb, analytes, headerlist, searchlist, max_row, max_col):
    return wb