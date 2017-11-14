from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import messagebox
from sys import exit
import csv
import tkinter as tk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from os.path import basename, splitext

##############################
# Global variables
root = tk.Tk()
max_row = 0
max_col = 0
num_samples = 0
##############################


def main():
    file, filename = get_file()
    wb = Workbook()
    wb = init_raw_data(file, wb)
    analytes = get_analytes(wb)
    wb = format_file(wb, analytes)
    save_file(file, wb, filename)


def get_file():
    global root
    root.withdraw()
    root.iconbitmap('proteinsimple_logo_bt.ico')
    success = False
    while not success:
        try:
            filename = askopenfilename(title='Choose your data files',
                                       multiple=False, filetypes=(('CSV Files', '*.csv'), ('All Files', '*.*')))
            if not filename:
                exit()
            elif not filename.endswith('.csv'):
                success = False
                messagebox.showerror(message="Invalid Filetype.",
                                     title="Failure")
            else:
                success = True
        except csv.Error as error:
            messagebox.showerror(message="Invalid Filetype.",
                                 title="Failure")

    file = open(filename)
    return file, filename


def save_file(file, wb, filename):
    newName = splitext(basename(filename))[0]
    options = {}
    options['defaultextension'] = ".xlsx"
    options['filetypes'] = (('xlsx files', '*.xlsx'), ('all files', '*.*'))
    #options['initialdir'] = ""
    options['initialfile'] = newName
    options['title'] = "Save as..."

    dest_filename = asksaveasfilename(**options)

    try:
        wb.save(filename=dest_filename)
    except PermissionError as e:
        messagebox.showerror(message="The file you are trying to overwrite is open. Close it and try again",
                             title="Failure")
        exit()

    if not dest_filename:
        exit()


def init_raw_data(file, wb):
    # Create first sheet
    global max_col
    global max_row
    csv.register_dialect('comma', delimiter=',')
    reader = csv.reader(file, dialect='comma')
    ws1 = wb.worksheets[0]
    ws1.title = 'Raw data'
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws1['%s%s' % (column_letter, (row_index + 1))].value = cell
            max_col = column_index + 1
        max_row = row_index + 1
    return wb


def get_analytes(wb):
    try:
        ws1 = wb.worksheets[0]
        analyte_order = []
        done = False
        count = 1
        while not done:
            letter = get_column_letter(count)
            ind = letter + '1'
            if ws1[ind].value == "AnalyteName":
                count = 2
                while not done:
                    ind = letter + str(count)
                    if not analyte_order:
                        analyte_order.append(str(ws1[ind].value))
                    else:
                        obj = str(ws1[ind].value)
                        if not obj in analyte_order:
                            analyte_order.append(obj)
                        else:
                            done = True
                    count += 1
            count += 1
            if count == max_col:
                raise ValueError('Item not found')
    except ValueError as error:
        messagebox.showerror(message="Missing Analyte Names.  Please export your data with "
                                     "this item included", title="Failure")
        exit()
    global num_samples
    num_samples = get_num_samples(ws1, analyte_order)
    return analyte_order


def format_file(wb, analytes):
    num_samples = get_num_samples(wb, analytes)
    if len(analytes) == 1:
        from one_by_72 import format
    else:
        if num_samples == 16:
            from four_by_16 import format
        elif num_samples == 32:
            from four_by_32 import format
    wb = format(wb, analytes, max_row, max_col)
    return wb


def get_num_samples(ws, analytes):
    return int((max_row - 1) / len(analytes))


if __name__ == '__main__':
    exit(main())
