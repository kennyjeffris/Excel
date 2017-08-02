"""
Script to format csv file.
Release version 1.0.1
Author: Kenny Jeffris
"""

##############################################
# Import libraries and functions
import sys
import csv
import string
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import ScatterChart, Reference, Series, marker
from openpyxl.drawing.fill import ColorChoice
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import messagebox
import tkinter as tk
import numpy as np

##############################################
# Open the file to format.
root = tk.Tk()
root.withdraw()
root.iconbitmap('proteinsimple_logo_bt.ico')
success = False
while not success:
    try:
        filename = askopenfilename(title='Choose your data files',
                                   multiple=False, filetypes=(('CSV Files', '*.csv'), ('All Files', '*.*')))
        if not filename:
            sys.exit()
        elif not filename.endswith('.csv'):
            success = False
            messagebox.showerror(message="Invalid Filetype.",
                                 title="Failure")
        else:
            success = True
    except csv.Error as error:
        messagebox.showerror(message="Invalid Filetype.",
                             title="Failure")

f = open(filename)

##############################################
# Create Style variable
medium = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

thin = Border(left=Side(style='thin'),
              right=Side(style='thin'),
              top=Side(style='thin'),
              bottom=Side(style='thin'))

medium_thinbottom = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

center_center = Alignment(horizontal='center', vertical='center')

yellow_fill = PatternFill('solid', fgColor=colors.YELLOW)
##############################################
# Helper functions


def col2num(colindex):
    """Convert excel column letter to index number."""
    number = 0
    for c in colindex:
        if c in string.ascii_letters:
            number = number * 26 + (ord(c.upper()) - ord('A')) + 1
    return number


def getitems(sheet, analyte, data):
    """Get the items in a column of a request sheet."""
    try:
        done = False
        count = 1
        while not done:
            letter = get_column_letter(count)
            ind = letter + '1'
            if data == sheet[ind].value:
                item = []
                for l in range(analyte, max_row, 4):
                    ind = letter + str(l + 1)
                    cell = sheet[ind]
                    try:
                        if cell.value.isspace() or cell.value == '' or cell.value is None or cell.value == "NaN":
                            item.append("ND")
                        else:
                            item.append(float(cell.value))
                    except Exception:
                        item.append(cell.value)
                return item
            count += 1
            if count > max_col:
                raise ValueError('Item not found')
    except ValueError as error:
        messagebox.showerror(message="Missing item {}.  Please export your data with "
                             "this item included".format(feature), title="Failure")
        sys.exit()


def as_text(value):
    """Return the input as a string."""
    if value is None:
        return ""
    return str(value)


def polyfit(x, coefficients):
    return (coefficients[3] + (coefficients[0] - coefficients[3]) /
            ((1 + (x / coefficients[2]) ** coefficients[1]) ** coefficients[4]))


##############################################
# Set up CSV tools
csv.register_dialect('comma', delimiter=',')
reader = csv.reader(f, dialect='comma')

# Initialize .xlsx file
wb = Workbook()

# Create first sheet
ws1 = wb.worksheets[0]
ws1.title = 'Raw data'

##############################################
# Copy Raw Data to sheet1
for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws1['%s%s' % (column_letter, (row_index + 1))].value = cell
        max_col = column_index + 1
    max_row = row_index + 1
##############################################

# Get the order of analytes and make a usable list
try:
    analyteOrder = []
    done = False
    count = 1
    while not done:
        letter = get_column_letter(count)
        ind = letter + '1'
        if ws1[ind].value == "AnalyteName":
            for l in range(2, 6):
                ind = letter + str(l)
                try:
                    analyteOrder.append(str(ws1[ind].value))
                except Exception:
                    analyteOrder.append(ws1[ind].value)
            break
        count += 1
        if count == max_col:
            raise ValueError('Item not found')
except ValueError as error:
    messagebox.showerror(message="Missing Analyte Names.  Please export your data with "
                                 "this item included", title="Failure")
    sys.exit()


# Lists for Summary 1
searchList1 = ['Gnr1Background', 'Gnr1RFU', 'Gnr2RFU', 'Gnr3RFU',
               'Signal', 'RFUPercentCV', 'Gnr1Signal',	'Gnr2Signal', 'Gnr3Signal',
               'RFU', 'Gnr1CalculatedConcentration',
               'Gnr2CalculatedConcentration',
               'Gnr3CalculatedConcentration', 'CalculatedConcentration',
               'CalculatedConcentrationPercentCV']

headerList1 = ['Sample', 'Bkgd', 'Gnr1', 'Gnr2', 'Gnr3', 'Avg', '% CV', 'Gnr1', 'Gnr2',
               'Gnr3', 'Avg', 'Gnr1', 'Gnr2', 'Gnr3', 'Avg', '% CV']

# Lists for Summary 2
searchList2 = ['Gnr1CalculatedConcentration',
               'Gnr2CalculatedConcentration', 'Gnr3CalculatedConcentration', 'CalculatedConcentration',
               'CalculatedConcentrationPercentCV']

headerList2 = ['Sample', 'Gnr1', 'Gnr2', 'Gnr3', 'Avg', '% CV']

# Lists for Summary 3
headerList3 = (analyteOrder[:])
headerList3.insert(0, 'Sample')

headerList4 = ['CalculatedConcentration', 'CurveCoefficientA',
               'CurveCoefficientB', 'CurveCoefficientC', 'CurveCoefficientD',
               'CurveCoefficientG']

headerList5 = ['Curve Coefficients', 'A', 'B', 'C', 'D', 'G']
##############################################
# Populate Summary 1 and Summary 2
ws2 = wb.create_sheet(title='Summary 1')
ws3 = wb.create_sheet(title='Summary 2')

for page in range(1, 3):
    if page == 1:
        workingSheet = ws2
        headerList = headerList1
        searchList = searchList1
    else:
        workingSheet = ws3
        headerList = headerList2
        searchList = searchList2
    for i in range(0, 4):
        startRow = i * 20 + 1
        analyteString = 'Analyte {} ({})'.format(i+1, analyteOrder[i])
        workingSheet['A{}'.format(startRow)] = analyteString
        if page == 1:
            workingSheet.merge_cells(start_row=startRow + 1, start_column=2, end_row=startRow + 1, end_column=7)
            cell = workingSheet['B{}'.format(startRow+1)]
            cell.value = 'RFU'
            cell.alignment = center_center
            cell.fill = yellow_fill

            workingSheet.merge_cells(start_row=startRow + 1, start_column=8, end_row=startRow + 1, end_column=11)
            cell = workingSheet['H{}'.format(startRow + 1)]
            cell.value = 'RFU-Bkgd'
            cell.alignment = center_center
            cell.fill = yellow_fill

            workingSheet.merge_cells(start_row=startRow + 1, start_column=12, end_row=startRow + 1, end_column=16)
            cell = workingSheet['L{}'.format(startRow + 1)]
            cell.value = 'Calculated Concentration'
            cell.alignment = center_center
            cell.fill = yellow_fill
            for index, col in enumerate(iterable=workingSheet.iter_cols(min_row=startRow+1, max_row=startRow+1,
                                                                        min_col=2, max_col=16)):
                for cell in col:
                    cell.border = medium_thinbottom
        else:
            workingSheet.merge_cells(start_row=startRow + 1, start_column=2, end_row=startRow + 1, end_column=6)
            cell = workingSheet['B{}'.format(startRow + 1)]
            cell.value = 'Calculated Concentration'
            cell.alignment = center_center
            cell.fill = yellow_fill
            for index, col in enumerate(iterable=workingSheet.iter_cols(min_row=startRow+1, max_row=startRow+1,
                                                                        min_col=2, max_col=6)):
                for cell in col:
                    cell.border = medium_thinbottom
        for index, col in enumerate(iterable=workingSheet.iter_cols(
                                        min_row=startRow+2,
                                        min_col=1, max_row=startRow+2,
                                        max_col=len(headerList))):
            for cell in col:
                cell.value = headerList[index]
                cell.border = medium
                if i == 0:
                    length = len(as_text(cell.value)) + 2
                    workingSheet.column_dimensions[cell.column].width = length

        for index, row in enumerate(iterable=workingSheet.iter_rows(
                                        min_row=startRow+3,
                                        max_col=1, max_row=startRow+18)):
            for cell in row:
                cell.value = index+1
                cell.border = medium_thinbottom

        for index, col in enumerate(iterable=workingSheet.iter_cols(
                                        min_col=2, max_col=len(headerList))):
            feature = searchList[index]
            values = getitems(ws1, i+1, feature)
            for index2, row in enumerate(iterable=workingSheet.iter_rows(
                                            min_col=index+2,
                                            min_row=startRow+3,
                                            max_row=startRow+18)):
                for cell in row:
                    cell.value = values[index2]
                    cell.border = thin
##############################################
# Populate Summary 3
ws4 = wb.create_sheet(title='Summary 3')
ws4['A1'].value = headerList4[0]
for index, col in enumerate(iterable=ws4.iter_cols(
                                min_row=2,
                                min_col=1, max_row=2,
                                max_col=len(headerList3))):
    for cell in col:
        cell.value = headerList3[index]
        cell.border = medium

for index, row in enumerate(iterable=ws4.iter_rows(
                                min_row=3,
                                max_col=1, max_row=18)):
    for cell in row:
        cell.value = index+1
        cell.border = medium_thinbottom

for index, col in enumerate(iterable=ws4.iter_cols(
                                min_col=2, max_col=len(headerList3),
                                min_row=3, max_row=18)):

    values = getitems(ws1, index+1, headerList4[0])
    for index2, row in enumerate(iterable=ws4.iter_rows(
                                    min_col=index+2,
                                    min_row=3,
                                    max_row=18)):
        for cell in row:
            cell.value = values[index2]
            cell.border = thin

ws4['A20'].value = headerList5[0]
ws4['A21'].border = medium
for index, col in enumerate(iterable=ws4.iter_cols(min_row=21,
                            min_col=2, max_row=21, max_col=5)):
    for cell in col:
        cell.value = headerList3[index + 1]
        cell.border = medium

for index, row in enumerate(iterable=ws4.iter_rows(min_row=22, max_row=26,
                                                   min_col=1, max_col=1)):
    for cell in row:
        cell.value = headerList5[index + 1]
        cell.border = medium_thinbottom

for index, col in enumerate(iterable=ws4.iter_cols(
                                min_col=2, max_col=len(headerList3),
                                min_row=22, max_row=26)):

    for index2, row in enumerate(iterable=ws4.iter_rows(
                                    min_col=index+2,
                                    min_row=22,
                                    max_row=26)):
        values = getitems(ws1, index+1, headerList4[index2+1])
        for cell in row:
            cell.value = values[0]
            cell.border = thin
##############################################
# Add plots to Summary 3

colors = [ColorChoice(prstClr="red"), ColorChoice(prstClr="orange"), ColorChoice(prstClr="blue"),
          ColorChoice(prstClr="green")]

for index, col in enumerate(iterable=ws4.iter_cols(
                                min_col=2,
                                max_col=5)):
    coeffs = []
    let = get_column_letter(2+index)
    for index2, row in enumerate(iterable=ws4.iter_rows(
                                    range_string='{}22:{}26'.format(let,
                                                                    let))):
        for cell in row:
            coeffs.append(cell.value)
    xvalues = []
    values = []
    xvalues.extend((.1, 1, 10, 100, 1000, 10000))
    for num in range(0, len(xvalues)):
        if isinstance(xvalues[num], (int, float, complex)):
            values.append(polyfit(xvalues[num], coeffs))
    for index3, row in enumerate(iterable=ws4.iter_rows(
                                    min_row=3,
                                    max_row=18,
                                    min_col=2+index,
                                    max_col=2+index)):
        for cell in row:
            xval = cell.value
            if isinstance(xval, (int, float, complex)) and not np.isnan(xval):
                xvalues.append(xval)
                values.append(polyfit(xval, coeffs))

    xvalues, values = zip(*sorted(zip(xvalues, values)))
    xvalues, values = (list(t) for t in zip(*sorted(zip(xvalues, values))))
    count1 = 0
    count2 = 0
    for index4, row in enumerate(iterable=ws4.iter_rows(
                            range_string='{}28:{}72'.format(let, let))):
        for cell in row:
            if count1 < len(xvalues):
                cell.value = xvalues[count1]
                count1 += 1
            elif count1 >= len(values) and count2 < len(values):
                cell.value = values[count2]
                count2 += 1
            else:
                break
    xref = Reference(ws4, min_col=2+index, max_col=2+index,
                     min_row=28, max_row=27+len(xvalues))
    yref = Reference(ws4, min_col=2+index, max_col=2+index,
                     min_row=28+len(xvalues),
                     max_row=27+len(xvalues)+len(values))

    chart = ScatterChart()
    chart.title = '{}'.format(headerList3[index+1])
    chart.style = 13
    chart.legend = None
    chart.x_axis.scaling.logBase = 10
    chart.y_axis.scaling.logBase = 10
    chart.x_axis.scaling.min = 0.1
    chart.y_axis.scaling.min = 0.1
    chart.x_axis.scaling.max = 10000
    chart.y_axis.scaling.max = 10000
    chart.y_axis.crossesAt = 0.1
    chart.y_axis.title = 'RFU'
    chart.x_axis.title = 'Concentration (pg/ml)'
    chart.x_axis.tickLblPos = "low"
    series = Series(yref, xref, title_from_data=False)
    series.marker = marker.Marker('circle')
    series.smooth = True
    series.graphicalProperties.line.width = 40000
    series.graphicalProperties.line.solidFill = colors[index]
    series.marker.graphicalProperties.solidFill = colors[index]  # Marker filling
    series.marker.graphicalProperties.line.solidFill = colors[index]  # Marker outline
    chart.series.append(series)
    ws4.add_chart(chart, 'H{}'.format((index * 15) + 1))

##############################################
# Save the resulting file
newName = 'output.xlsx'

dest_filename = asksaveasfilename(title='Save File.', filetypes=(('xlsx files', '*.xlsx'), ('all files', '*.*')),                                  initialfile=newName)
try:
    wb.save(filename=dest_filename)
except PermissionError as e:
    messagebox.showerror(message="The file you are trying to overwrite is open. Close it and try again",
                         title="Failure")
    sys.exit()

if not dest_filename:
    sys.exit()

