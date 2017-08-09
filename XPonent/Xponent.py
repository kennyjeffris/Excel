"""
Script to format Xponent output .csv file.
Release version 1.0
Author: Kenny Jeffris
"""

##############################################
# Import libraries and functions
import sys
import csv
import string
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import ScatterChart, Reference, Series, marker
from openpyxl.drawing.fill import ColorChoice
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import messagebox
import tkinter as tk

##############################################
# Open the file to format.
root = tk.Tk()
root.withdraw()
root.iconbitmap('proteinsimple_logo_bt.ico')
success = False
while not success:
    try:
        filename = askopenfilename(title='Choose your data files',
                                   multiple=False, filetypes=(('CSV Files', '*.csv'), ('All Files', '*.*')))
        if not filename:
            sys.exit()
        elif not filename.endswith('.csv'):
            success = False
            messagebox.showerror(message="Invalid Filetype.",
                                 title="Failure")
        else:
            success = True
    except csv.Error as error:
        messagebox.showerror(message="Invalid Filetype.",
                             title="Failure")

f = open(filename)

##############################################
# Create Style variable
medium = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

thin = Border(left=Side(style='thin'),
              right=Side(style='thin'),
              top=Side(style='thin'),
              bottom=Side(style='thin'))

medium_thinbottom = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

center_center = Alignment(horizontal='center', vertical='center')

right_center = Alignment(horizontal='right', vertical='center')

yellow_fill = PatternFill('solid', fgColor=colors.YELLOW)
red_fill = PatternFill('solid', fgColor=colors.RED)
##############################################
# Helper functions


def col2num(colindex):
    """Convert excel column letter to index number."""
    number = 0
    for c in colindex:
        if c in string.ascii_letters:
            number = number * 26 + (ord(c.upper()) - ord('A')) + 1
    return number


def get_items(sheet, analyte, data):
    """Get the items in a column of a request sheet."""
    try:
        done = False
        count = 1
        while not done:
            letter = get_column_letter(count)
            ind = letter + '1'
            if data == sheet[ind].value:
                item = []
                for l in range(analyte, max_row, 4):
                    ind = letter + str(l + 1)
                    cell = sheet[ind]
                    try:
                        if cell.value.isspace() or cell.value == '' or cell.value is None or cell.value == "NaN":
                            item.append("ND")
                        else:
                            item.append(float(cell.value))
                    except Exception:
                        item.append(cell.value)
                return item
            count += 1
            if count > max_col:
                raise ValueError('Item not found')
    except ValueError as error:
        messagebox.showerror(message="Missing item {}.  Please export your data with "
                             "this item included".format(feature), title="Failure")
        sys.exit()


def as_text(value):
    """Return the input as a string."""
    if value is None:
        return ""
    return str(value)


def poly_fit(x, coefficients):
    return (coefficients[3] + (coefficients[0] - coefficients[3]) /
            ((1 + (x / coefficients[2]) ** coefficients[1]) ** coefficients[4]))


##############################################
# Set up CSV tools
csv.register_dialect('comma', delimiter=',')
reader = csv.reader(f, dialect='comma')

# Initialize .xlsx file
wb = Workbook()

# Create first sheet
ws1 = wb.worksheets[0]
ws1.title = 'Raw data'

##############################################
# Copy Raw Data to sheet1
for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws1['%s%s' % (column_letter, (row_index + 1))].value = cell
        max_col = column_index + 1
    max_row = row_index + 1

##############################################
# Save the resulting file
newName = 'output.xlsx'

dest_filename = asksaveasfilename(title='Save File.', filetypes=(('xlsx files', '*.xlsx'), ('all files', '*.*')),                                  initialfile=newName)
try:
    wb.save(filename=dest_filename)
except PermissionError as e:
    messagebox.showerror(message="The file you are trying to overwrite is open. Close it and try again",
                         title="Failure")
    sys.exit()

if not dest_filename:
    sys.exit()


