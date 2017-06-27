"""
Script to format csv file.

Author: Kenny Jeffris
"""

##############################################
# Import libraries and functions
import csv
import string
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import LineChart, Reference
# from os.path import split, join
import easygui
##############################################
# Open the file to format.
filename = easygui.fileopenbox(msg='Choose your data file.',
                               filetypes=['*.csv'])
# filename = 'C:\\Users\\kjeffris\\My Documents\\Excel\\export.csv'
f = open(filename)
# Change this when implementing into program
##############################################
# Create Style variable
medium = Border(left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

thin = Border(left=Side(style='thin'),
              right=Side(style='thin'),
              top=Side(style='thin'),
              bottom=Side(style='thin'))

medium_thinbottom = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

##############################################
# Helper functions


def col2num(col):
    """Convert excel column letter to index number."""
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def getItems(sheet, analyte, feature):
    """Get the items in a column of a request sheet."""
    done = False
    count = 1
    while (not done):
        column_letter = get_column_letter(count)
        index = column_letter + '1'
        if feature == sheet[index].value:
            item = []
            for i in range(analyte, 65, 4):
                index = column_letter + str(i + 1)
                try:
                    item.append(float(sheet[index].value))
                except Exception:
                    item.append('')
            done = True
            return item
        count += 1
    return []


def as_text(value):
    """Return the input as a string."""
    if value is None:
        return ""
    return str(value)


def createChart(title, style, ytitle, xtitle, sheet, min_row, max_row, min_col,
                max_col):
    """Create a line chart."""
    chart = LineChart(title=title, style=style)
    chart.y_axis.title = ytitle
    chart.x_axis.title = xtitle
    data = Reference(sheet, min_col=min_col, min_row=min_row, max_col=max_col,
                     max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    return chart


##############################################
# Set up CSV tools
csv.register_dialect('comma', delimiter=',')
reader = csv.reader(f, dialect='comma')

# Initialize .xlsx file
wb = Workbook()
# (new, extra) = split(filename)
newName = 'output.xlsx'
# dest_filename = join(new, newName)
dest_filename = easygui.filesavebox(msg='Save File.', default='output.xlsx',
                                    filetypes=['*.xlsx'])
# Create first sheet
ws1 = wb.worksheets[0]
ws1.title = 'Raw data'

##############################################
# Copy Raw Data to sheet1
for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws1['%s%s' % (column_letter, (row_index + 1))].value = cell
##############################################
analyteOrder = []
for index, row in enumerate(iterable=ws1.iter_rows(min_row=2, max_row=5,
                                                   max_col=1)):
    for cell in row:
        analyteOrder.append(cell.value)

headerList1 = ['Sample', 'Gnr1Background', 'Gnr1RFU', 'Gnr2RFU', 'Gnr3RFU',
               'RFU', 'RFUPercentCV', 'Gnr1Signal',	'Gnr2Signal', 'Gnr3Signal',
               'Signal', 'Gnr1CalculatedConcentration',
               'Gnr2CalculatedConcentration',
               'Gnr3CalculatedConcentration', 'CalculatedConcentration',
               'CalculatedConcentrationPercentCV']

headerList2 = ['Sample', 'Gnr1CalculatedConcentration',
               'Gnr2CalculatedConcentration', 'Gnr3CalculatedConcentration',
               'CalculatedConcentrationPercentCV']

headerList3 = (analyteOrder[:])
headerList3.insert(0, 'Sample')

headerList4 = ['CalculatedConcentration', 'CurveCoefficientA',
               'CurveCoefficientB', 'CurveCoefficientC', 'CurveCoefficientD',
               'CurveCoefficientG']

headerList5 = ['Curve Coefficients', 'A', 'B', 'C', 'D', 'G']
##############################################
# Populate Summary 1 and Summary 2
ws2 = wb.create_sheet(title='Summary 1')
ws3 = wb.create_sheet(title='Summary 2')

for page in range(1, 3):
    if (page == 1):
        workingSheet = ws2
        headerList = headerList1
    else:
        workingSheet = ws3
        headerList = headerList2
    for i in range(0, 4):
        analyteString = 'Analyte {} ({})'.format(i+1, analyteOrder[i])
        startRow = i*20+1
        workingSheet['A{}'.format(startRow)] = analyteString
        for index, col in enumerate(iterable=workingSheet.iter_cols(
                                        min_row=startRow+1,
                                        min_col=1, max_row=startRow+1,
                                        max_col=len(headerList))):
            for cell in col:
                cell.value = headerList[index]
                cell.border = medium
                if (i == 0):
                    length = len(as_text(cell.value)) + 2
                    workingSheet.column_dimensions[cell.column].width = length

        for index, row in enumerate(iterable=workingSheet.iter_rows(
                                        min_row=startRow+2,
                                        max_col=1, max_row=startRow+17)):
            for cell in row:
                cell.value = index+1
                cell.border = medium_thinbottom

        for index, col in enumerate(iterable=workingSheet.iter_cols(
                                        min_col=2, max_col=len(headerList))):
            feature = headerList[index + 1]
            values = getItems(ws1, i+1, feature)
            for index2, row in enumerate(iterable=workingSheet.iter_rows(
                                            min_col=index+2,
                                            min_row=startRow+2,
                                            max_row=startRow+17)):
                for cell in row:
                    cell.value = values[index2]
                    cell.border = thin
##############################################
# Populate Summary 3
ws4 = wb.create_sheet(title='Summary 3')
ws4['A1'].value = headerList4[0]
for index, col in enumerate(iterable=ws4.iter_cols(
                                min_row=2,
                                min_col=1, max_row=2,
                                max_col=len(headerList3))):
    for cell in col:
        cell.value = headerList3[index]
        cell.border = medium

for index, row in enumerate(iterable=ws4.iter_rows(
                                min_row=3,
                                max_col=1, max_row=18)):
    for cell in row:
        cell.value = index+1
        cell.border = medium_thinbottom

for index, col in enumerate(iterable=ws4.iter_cols(
                                min_col=2, max_col=len(headerList3),
                                min_row=3, max_row=18)):

    values = getItems(ws1, index+1, headerList4[0])
    for index2, row in enumerate(iterable=ws4.iter_rows(
                                    min_col=index+2,
                                    min_row=3,
                                    max_row=18)):
        for cell in row:
            cell.value = values[index2]
            cell.border = thin

ws4['A20'].value = headerList5[0]
ws4['A21'].border = medium
for index, col in enumerate(iterable=ws4.iter_cols(min_row=21,
                            min_col=2, max_row=21, max_col=5)):
    for cell in col:
        cell.value = headerList3[index + 1]
        cell.border = medium

for index, row in enumerate(iterable=ws4.iter_rows(min_row=22, max_row=26,
                                                   min_col=1, max_col=1)):
    for cell in row:
        cell.value = headerList5[index + 1]
        cell.border = medium_thinbottom

for index, col in enumerate(iterable=ws4.iter_cols(
                                min_col=2, max_col=len(headerList3),
                                min_row=22, max_row=26)):

    for index2, row in enumerate(iterable=ws4.iter_rows(
                                    min_col=index+2,
                                    min_row=22,
                                    max_row=26)):
        values = getItems(ws1, index+1, headerList4[index2+1])
        for cell in row:
            cell.value = values[0]
            cell.border = thin
##############################################
# Save the resulting file
wb.save(filename=dest_filename)
